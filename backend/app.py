# backend/app.py
import sys
import json
import re
from pathlib import Path
from PyPDF2 import PdfReader
from openpyxl import Workbook
from datetime import datetime

def load_config(config_path):
    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def clean_number(text):
    text = text.replace('.', '').replace(',', '.')
    if re.match(r"^\(\d+\.\d+\)$", text.strip()):
        return '-' + text.strip()[1:-1]
    return text.strip()

def extract_data(text, config):
    results = {}
    for key, spec in config['keywords'].items():
        mode = spec['mode']
        if mode == 'two_numbers_after':
            match = re.search(re.escape(key) + r'[^0-9\-]+([\(]?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})[\)]?)\D+([\(]?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})[\)]?)', text)
            if match:
                results[key] = [clean_number(match.group(1)), clean_number(match.group(2))]
            else:
                results[key] = ["", ""]
        elif mode == 'until_dot':
            match = re.search(re.escape(key) + r'(.+?)\.', text)
            results[key] = match.group(1).strip() if match else ""
        elif mode == 'until_newline':
            match = re.search(re.escape(key) + r'([^\n\r]+)', text)
            results[key] = match.group(1).strip() if match else ""
        elif mode == 'between_phrases':
            start = spec['start_phrase']
            end = spec['end_phrase']
            match = re.search(re.escape(start) + r'(.*?)' + re.escape(end), text, re.DOTALL)
            if match:
                # Elimina saltos de línea y espacios extra
                value = match.group(1).replace('\n', ' ').replace('\r', ' ').strip()
                value = re.sub(r'\s+', ' ', value)  # Opcional: reemplaza múltiples espacios por uno solo
                results[key] = value
            else:
                results[key] = ""
        elif mode == 'sum_percent_until_100':
            # Busca todos los países y sus porcentajes después de la clave
            bloque = ""
            bloque_match = re.search(re.escape(key) + r'(.*)', text, re.DOTALL | re.IGNORECASE)
            if bloque_match:
                bloque = bloque_match.group(1)
            else:
                bloque = text
            pattern = r'([A-Za-zÁÉÍÓÚÜÑáéíóúüñ\s]+)\s*Percentage of turnover\s*:\s*(\d+)%'
            matches = re.findall(pattern, bloque, re.IGNORECASE)
            suma = 0
            resultado = []
            for pais, porcentaje in matches:
                porcentaje = int(porcentaje)
                suma += porcentaje
                resultado.append(f"{pais.strip()}: {porcentaje}%")
                if suma >= 100:
                    break
            results[key] = resultado
    return results

def save_to_excel(data, output_path, config=None):
    wb = Workbook()
    ws = wb.active

    # --- VALORES ESTÁTICOS ---
    ws['A1'] = "Reporte generado"
    ws['B1'] = "HOLA"
    ws['C1'] = "Mi Empresa S.A."

    # --- Tus datos dinámicos ---
    row = 3  # Por si quieres seguir mostrando todo en filas
    for key, value in data.items():
        # Si tienes config y la clave tiene definida una celda, úsala
        cell = None
        if config and key in config['keywords'] and 'cells' in config['keywords'][key]:
            cell = config['keywords'][key]['cells'][0]
        if cell:
            # Une la lista en una sola celda si es necesario
            if isinstance(value, list):
                ws[cell] = " - ".join(value)
            else:
                ws[cell] = value
        else:
            # Fallback: sigue escribiendo en filas si no hay celda definida
            ws.cell(row=row, column=1, value=key)
            if isinstance(value, list):
                ws.cell(row=row, column=2, value=" - ".join(value))
            else:
                ws.cell(row=row, column=2, value=value)
            row += 1

    wb.save(output_path)

def main():
    if len(sys.argv) < 4:
        print(json.dumps({'error': 'Faltan argumentos: pdf_path config_path output_path'}))
        return

    file_path = sys.argv[1]
    config_path = sys.argv[2]
    output_path = sys.argv[3]

    try:
        reader = PdfReader(file_path)
        full_text = "\n".join([page.extract_text() for page in reader.pages if page.extract_text()])
        config = load_config(config_path)
        extracted = extract_data(full_text, config)
        save_to_excel(extracted, output_path, config)
    except Exception as e:
        error_path = output_path.replace('.xlsx', '.error.json')
        with open(error_path, 'w', encoding='utf-8') as f:
            json.dump({'error': str(e)}, f, ensure_ascii=False)

if __name__ == '__main__':
    main()
